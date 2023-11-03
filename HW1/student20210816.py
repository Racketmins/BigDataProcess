#!/user/bin/python3
from openpyxl import load_workbook

def countStudent(fn, cnt):
	try:
		wb = load_workbook(filename = fn)
		ws = wb.active
		row = 2
		while ws.cell(row = row, column = 1).value != None:
			cnt  += 1
			row += 1
		return cnt
	except FileNotFoundError:
		print("File not found-first")
	finally:
		wb.close()

def calTotal(fn, cnt):
	try:
		wb = load_workbook(filename = fn)
		ws = wb.active
		for row in range(2, cnt + 2):
			mid = ws.cell(row = row, column = 3).value * 0.3
			fin = ws.cell(row = row, column = 4).value * 0.35
			hw = ws.cell(row = row, column = 5).value * 0.34
			att = ws.cell(row = row, column = 6).value
			total = mid + fin + hw + att
			ws.cell(row = row, column = 7, value = total)
		wb.save(filename = fn)
	except FileNotFoundError:
		print("We can't find that file")
	finally:
		wb.close()

def rank(fn, cnt):
	try:
		wb = load_workbook(filename = fn)
		ws = wb.active
		score = [ws.cell(row = row, column = 7).value for row in range(2, cnt + 2)]
		score.sort()
		score.reverse()
					
		A = int(cnt * 0.3) - 1
		Ap = int((A + 1) * 0.5) -1
		B = int(cnt * 0.7) - 1
		Bp = int((B - A) * 0.5) + A
		F = 0
		for i in range(2, cnt + 2):
			if ws.cell(row = i, column = 7).value < 40:
				ws.cell(row = i, column = 8, value = 'F')
				F -= 1
		Cp = int((cnt - 1 + F - B) * 0.5) + B

		for i in range(2, cnt + 2):
			flag = ws.cell(row = i, column = 7).value
			if ws.cell(row = i, column = 7).value >= 40:
				if flag >= score[Ap]:
					ws.cell(row = i, column = 8, value = 'A+')
				elif flag >= score[A]:
					ws.cell(row = i, column = 8, value = 'A0')
				elif flag >= score[Bp]:
					ws.cell(row = i, column = 8, value = 'B+')
				elif flag >= score[B]:
					ws.cell(row = i, column = 8, value = 'B0')
				elif flag >= score[Cp]:
					ws.cell(row = i, column = 8, value = 'C+')
				else:
					ws.cell(row = i, column = 8, value = 'C0')
		wb.save(filename = fn)
	except FileNotFoundError:
		print("Not File")
	finally:
		wb.close()

count = 0
fileName = "student.xlsx"
count = countStudent(fileName, count)
calTotal(fileName, count)
rank(fileName, count)
